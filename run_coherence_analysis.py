"""
Spectral Coherence Analysis — MALES vs FEMALES
================================================
Replicates the exact analysis pipeline used in the Tracy Fiber Photometry
Processing Suite (static coherence, sliding coherence, rolling correlation).

Brain regions: BNST and CEA (channel 1 and channel 2)
Conditions:    PRE and POST
Sexes:         MALES and FEMALES

Outputs (saved to ./Coherence_Output/):
  - coherence_summary.xlsx         : mean coherence & correlation per subject/condition
  - coherence_frequency_data.csv   : per-subject frequency-domain coherence values
  - coherence_rolling_data.csv     : per-subject rolling-correlation time series
  - figure_01_static_coherence.png : mean ± SEM spectral coherence (all 4 groups)
  - figure_02_sliding_heatmaps.png : group-averaged time-frequency coherence
  - figure_03_rolling_correlation.png : rolling window Pearson correlation
  - figure_04_summary_bars.png     : bar chart of mean coherence by group
"""

import os
import csv
import math
import numpy as np
import openpyxl
from openpyxl import Workbook
from scipy.signal import coherence as scipy_coherence
from scipy import stats
import matplotlib
matplotlib.use('Agg')  # Non-interactive backend — no display required
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib.patches import Patch

# ─────────────────────────────────────────────────────────────────────────────
# PATHS
# ─────────────────────────────────────────────────────────────────────────────
MALES_XLSX   = r'C:\Users\jober\Downloads\MALES_coherence.xlsx\MALES_coherence.xlsx'
FEMALES_XLSX = r'C:\Users\jober\Downloads\FEMALES_coherence.xlsx\FEMALES_coherence.xlsx'

SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
OUT_DIR      = os.path.join(SCRIPT_DIR, 'Coherence_Output')
os.makedirs(OUT_DIR, exist_ok=True)

# ─────────────────────────────────────────────────────────────────────────────
# ANALYSIS PARAMETERS  (Long Session preset — matches Tracy defaults)
# ─────────────────────────────────────────────────────────────────────────────
FMIN           = 0.05    # Hz  — low-frequency cutoff
FMAX           = 5.0     # Hz  — high-frequency cutoff
NPERSEG_SEC    = 20      # s   — Welch window
NOVERLAP_FRAC  = 0.5     # 50 % overlap
WINDOW_TYPE    = 'hann'

SLIDE_WIN_SEC  = 60      # s   — sliding coherence window
SLIDE_STEP_SEC = 5       # s   — sliding coherence step
SLIDE_NPERSEG  = 20      # s   — Welch window inside each sliding window

ROLL_WIN_SEC   = 60      # s   — rolling correlation window
ROLL_STEP_SEC  = 2       # s   — rolling correlation step
ROLL_MIN_FRAC  = 0.90    # minimum valid-sample fraction per window

# ─────────────────────────────────────────────────────────────────────────────
# PLOT STYLE  (publication-quality, matches Tracy's aesthetic)
# ─────────────────────────────────────────────────────────────────────────────
plt.rcParams.update({
    'font.family':      'DejaVu Sans',
    'font.size':        11,
    'axes.linewidth':   1.2,
    'axes.grid':        True,
    'grid.alpha':       0.3,
    'axes.spines.top':  False,
    'axes.spines.right':False,
    'figure.dpi':       150,
    'savefig.dpi':      300,
    'savefig.bbox':     'tight',
})

# Group colour palette ────────────────────────────────────────────────────────
COLOURS = {
    'Males PRE':    '#5B8DB8',   # steel blue
    'Males POST':   '#1A3F6F',   # dark navy
    'Females PRE':  '#D98080',   # dusty rose
    'Females POST': '#8B0000',   # dark red
}
SHADE_ALPHA = 0.20   # SEM shading

# ─────────────────────────────────────────────────────────────────────────────
# DATA LOADING
# ─────────────────────────────────────────────────────────────────────────────

def load_sheet(wb, sheet_name):
    """Return (subjects, time_vec, data_array) from a single xlsx sheet.

    data_array shape: (n_samples, n_subjects)
    Subjects are read from the first row; TIME is the last non-None column.
    """
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]

    # Find TIME column (last non-None header entry)
    time_col = None
    for idx, h in enumerate(header):
        if h == 'TIME':
            time_col = idx
            break

    # Subject columns (all non-None, non-TIME header entries)
    subj_cols = [(idx, str(h)) for idx, h in enumerate(header)
                 if h is not None and h != 'TIME']

    n_rows = len(rows) - 1
    subjects = [sc[1] for sc in subj_cols]
    data = np.full((n_rows, len(subjects)), np.nan)
    time_vec = np.full(n_rows, np.nan)

    for row_i, row in enumerate(rows[1:]):
        if time_col is not None and row[time_col] is not None:
            time_vec[row_i] = float(row[time_col])
        for col_i, (col_idx, _) in enumerate(subj_cols):
            val = row[col_idx]
            if val is not None:
                data[row_i, col_i] = float(val)

    return subjects, time_vec, data


def load_sex_data(xlsx_path):
    """Load all four sheets from one sex's xlsx file.

    Returns dict:
        {
            'PRE':  {'BNST': (subjects, time, data), 'CEA': (subjects, time, data)},
            'POST': {'BNST': (subjects, time, data), 'CEA': (subjects, time, data)},
        }
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    out = {}
    for cond in ('PRE', 'POST'):
        out[cond] = {}
        for region in ('BNST', 'CEA'):
            sheet = f'{cond} {region}'
            subjects, time_vec, data = load_sheet(wb, sheet)
            out[cond][region] = (subjects, time_vec, data)
    wb.close()
    return out

# ─────────────────────────────────────────────────────────────────────────────
# ANALYSIS FUNCTIONS  (exact replication of Tracy methods)
# ─────────────────────────────────────────────────────────────────────────────

def infer_fs(time_vec):
    """Estimate sampling frequency from a time vector."""
    dts = np.diff(time_vec[~np.isnan(time_vec)])
    return 1.0 / float(np.median(dts))


def calculate_static_coherence(s1, s2, fs, fmin=FMIN, fmax=FMAX,
                                nperseg_sec=NPERSEG_SEC):
    """Welch-based spectral coherence between s1 and s2.

    Matches Tracy's calculate_static_coherence:
      scipy.signal.coherence, hann window, 50 % overlap.
    Returns (freqs, coh) filtered to [fmin, fmax], or (None, None) on failure.
    """
    s1 = np.asarray(s1, dtype=float).ravel()
    s2 = np.asarray(s2, dtype=float).ravel()
    valid = ~(np.isnan(s1) | np.isnan(s2))
    s1, s2 = s1[valid], s2[valid]

    nperseg = int(nperseg_sec * fs)
    noverlap = int(nperseg * NOVERLAP_FRAC)

    if len(s1) < nperseg * 2:
        return None, None

    freqs, coh = scipy_coherence(s1, s2, fs=fs, nperseg=nperseg,
                                  noverlap=noverlap, window=WINDOW_TYPE)
    mask = (freqs >= fmin) & (freqs <= fmax)
    return freqs[mask], coh[mask]


def calculate_sliding_coherence(s1, s2, fs,
                                 fmin=FMIN, fmax=FMAX,
                                 win_sec=SLIDE_WIN_SEC,
                                 step_sec=SLIDE_STEP_SEC,
                                 nperseg_sec=SLIDE_NPERSEG):
    """Sliding-window spectral coherence.

    Matches Tracy's calculate_sliding_coherence.
    Returns (times, freqs, coh_tf) where coh_tf is (n_freqs, n_times),
    or (None, None, None) on failure.
    """
    s1 = np.asarray(s1, dtype=float).ravel()
    s2 = np.asarray(s2, dtype=float).ravel()
    valid = ~(np.isnan(s1) | np.isnan(s2))
    s1, s2 = s1[valid], s2[valid]

    win     = int(win_sec  * fs)
    step    = int(step_sec * fs)
    nperseg = int(nperseg_sec * fs)
    noverlap = int(nperseg * NOVERLAP_FRAC)

    if len(s1) < win:
        return None, None, None

    times, coh_matrix, freqs = [], [], None

    for start in range(0, len(s1) - win + 1, step):
        w1 = s1[start: start + win]
        w2 = s2[start: start + win]
        if len(w1) >= nperseg * 2:
            f, c = scipy_coherence(w1, w2, fs=fs, nperseg=nperseg,
                                    noverlap=noverlap, window=WINDOW_TYPE)
            mask = (f >= fmin) & (f <= fmax)
            if freqs is None:
                freqs = f[mask]
            coh_matrix.append(c[mask])
            times.append(start / fs)

    if not coh_matrix:
        return None, None, None

    return np.array(times), freqs, np.array(coh_matrix).T  # (n_freqs, n_times)


def calculate_rolling_correlation(s1, s2, fs,
                                   win_sec=ROLL_WIN_SEC,
                                   step_sec=ROLL_STEP_SEC,
                                   min_valid_fraction=ROLL_MIN_FRAC):
    """Rolling Pearson correlation with Fisher-Z transform.

    Matches Tracy's calculate_rolling_correlation.
    Returns (times, r_values) or (None, None) on failure.
    """
    s1 = np.asarray(s1, dtype=float).ravel()
    s2 = np.asarray(s2, dtype=float).ravel()

    win  = int(win_sec  * fs)
    step = int(step_sec * fs)
    min_valid = int(win * min_valid_fraction)

    if len(s1) < win:
        return None, None

    times, r_vals = [], []

    for start in range(0, len(s1) - win + 1, step):
        w1 = s1[start: start + win]
        w2 = s2[start: start + win]
        valid = ~(np.isnan(w1) | np.isnan(w2))
        if valid.sum() >= min_valid:
            r = np.corrcoef(w1[valid], w2[valid])[0, 1]
            if not math.isnan(r):
                # Fisher Z-transform (clipped to avoid log(0))
                r_z = 0.5 * math.log((1 + r) / max(1 - r, 1e-10))
                r_vals.append(r_z)
                times.append(start / fs)

    if not r_vals:
        return None, None

    return np.array(times), np.array(r_vals)

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def mean_sem(arrays):
    """Return (mean, sem) across a list of equal-length 1-D arrays."""
    mat = np.array(arrays)
    m   = np.nanmean(mat, axis=0)
    sem = np.nanstd(mat, axis=0, ddof=1) / np.sqrt((~np.isnan(mat)).sum(axis=0))
    return m, sem


def pad_to_length(arr, length, fill=np.nan):
    """Pad or truncate a 1-D array to exactly `length` samples."""
    out = np.full(length, fill)
    n   = min(len(arr), length)
    out[:n] = arr[:n]
    return out

# ─────────────────────────────────────────────────────────────────────────────
# MAIN ANALYSIS LOOP
# ─────────────────────────────────────────────────────────────────────────────

def run_analysis():
    print("Loading data…")
    sex_data = {
        'Males':   load_sex_data(MALES_XLSX),
        'Females': load_sex_data(FEMALES_XLSX),
    }

    # Storage for all results
    results = {}   # key: e.g. 'Males PRE', value: dict of lists

    for sex, sex_dict in sex_data.items():
        for cond, cond_dict in sex_dict.items():
            group_key = f'{sex} {cond}'
            print(f"  Processing {group_key}…")

            subjects  = cond_dict['BNST'][0]
            time_vec  = cond_dict['BNST'][1]
            bnst_data = cond_dict['BNST'][2]   # (n_samples, n_subjects)
            cea_data  = cond_dict['CEA'][2]

            fs = infer_fs(time_vec)
            print(f"    fs = {fs:.2f} Hz, {len(subjects)} subjects, "
                  f"{bnst_data.shape[0]} samples")

            group_results = {
                'subjects':          subjects,
                'fs':                fs,
                'time_vec':          time_vec,
                # per-subject lists — filled below
                'static_freqs':      None,   # shared across subjects
                'static_coh':        [],     # list of 1-D arrays
                'sliding_times':     None,
                'sliding_freqs':     None,
                'sliding_coh':       [],     # list of (n_freqs, n_times) arrays
                'rolling_times':     None,
                'rolling_r':         [],
                'mean_coherence':    [],
                'mean_correlation':  [],
            }

            for i, subj in enumerate(subjects):
                s1 = bnst_data[:, i]
                s2 = cea_data[:, i]

                # Static coherence
                f, c = calculate_static_coherence(s1, s2, fs)
                if c is not None:
                    if group_results['static_freqs'] is None:
                        group_results['static_freqs'] = f
                    group_results['static_coh'].append(c)
                    group_results['mean_coherence'].append(float(np.mean(c)))

                # Sliding coherence
                t_sl, f_sl, coh_tf = calculate_sliding_coherence(s1, s2, fs)
                if coh_tf is not None:
                    if group_results['sliding_times'] is None:
                        group_results['sliding_times'] = t_sl
                        group_results['sliding_freqs'] = f_sl
                    # Align to shared time grid
                    n_t = len(group_results['sliding_times'])
                    n_f = len(group_results['sliding_freqs'])
                    aligned = np.full((n_f, n_t), np.nan)
                    nt_use = min(n_t, coh_tf.shape[1])
                    aligned[:, :nt_use] = coh_tf[:, :nt_use]
                    group_results['sliding_coh'].append(aligned)

                # Rolling correlation
                t_r, r = calculate_rolling_correlation(s1, s2, fs)
                if r is not None:
                    if group_results['rolling_times'] is None:
                        group_results['rolling_times'] = t_r
                    n_t = len(group_results['rolling_times'])
                    group_results['rolling_r'].append(pad_to_length(r, n_t))
                    group_results['mean_correlation'].append(float(np.mean(np.abs(r))))

                print(f"      {subj}: mean_coh={group_results['mean_coherence'][-1]:.3f}"
                      f" | mean_|r|={group_results['mean_correlation'][-1]:.3f}"
                      if group_results['mean_coherence'] and group_results['mean_correlation']
                      else f"      {subj}: processed")

            results[group_key] = group_results

    print("\nGenerating figures…")
    plot_static_coherence(results)
    plot_sliding_heatmaps(results)
    plot_rolling_correlation(results)
    plot_summary_bars(results)

    print("\nExporting data files…")
    export_summary_xlsx(results)
    export_frequency_csv(results)
    export_rolling_csv(results)

    print(f"\nAll outputs saved to:\n  {OUT_DIR}")

# ─────────────────────────────────────────────────────────────────────────────
# FIGURES
# ─────────────────────────────────────────────────────────────────────────────

def _adj_time(time_arr):
    """Shift a time array so that the first sample aligns with the original
    recording time offset (i.e., the raw seconds from the xlsx TIME column).
    The sliding/rolling times are 0-based sample indices; add the first
    recorded timestamp to convert back to session time."""
    return time_arr   # already in sample seconds from start of signal


def plot_static_coherence(results):
    """Figure 1 — static spectral coherence, all groups overlay."""
    fig, axes = plt.subplots(1, 2, figsize=(13, 5), sharey=True)
    fig.suptitle('Spectral Coherence: BNST–CEA\n(Welch method, Hann window, '
                 f'{NPERSEG_SEC} s segments, {FMIN}–{FMAX} Hz)',
                 fontsize=13, fontweight='bold')

    panels = {
        axes[0]: ['Males PRE',   'Males POST'],
        axes[1]: ['Females PRE', 'Females POST'],
    }

    for ax, group_keys in panels.items():
        title_sex = group_keys[0].split()[0]
        ax.set_title(f'{title_sex}', fontsize=12)
        for gk in group_keys:
            if gk not in results or not results[gk]['static_coh']:
                continue
            freqs = results[gk]['static_freqs']
            coh_list = results[gk]['static_coh']
            colour = COLOURS[gk]

            # Individual subject traces
            for c in coh_list:
                ax.plot(freqs, c, color=colour, linewidth=0.8, alpha=0.35)

            # Mean ± SEM
            m, sem = mean_sem(coh_list)
            ax.plot(freqs, m, color=colour, linewidth=2.2, label=gk)
            ax.fill_between(freqs, m - sem, m + sem,
                            color=colour, alpha=SHADE_ALPHA)

        ax.set_xlabel('Frequency (Hz)')
        ax.set_ylabel('Coherence')
        ax.set_ylim([0, 1])
        ax.legend(framealpha=0.8, fontsize=9)

    fig.tight_layout()
    path = os.path.join(OUT_DIR, 'figure_01_static_coherence.png')
    fig.savefig(path)
    plt.close(fig)
    print(f"  Saved: {path}")


def plot_sliding_heatmaps(results):
    """Figure 2 — group-averaged sliding-window coherence heatmaps."""
    group_keys = ['Males PRE', 'Males POST', 'Females PRE', 'Females POST']
    valid_keys = [gk for gk in group_keys
                  if gk in results and results[gk]['sliding_coh']]

    if not valid_keys:
        print("  No sliding coherence data to plot.")
        return

    n = len(valid_keys)
    fig, axes = plt.subplots(1, n, figsize=(5 * n, 5), sharey=True)
    if n == 1:
        axes = [axes]
    fig.suptitle('Time–Frequency Coherence: BNST–CEA\n'
                 f'(sliding window {SLIDE_WIN_SEC} s, step {SLIDE_STEP_SEC} s)',
                 fontsize=13, fontweight='bold')

    vmin, vmax = 0.0, 1.0

    for ax, gk in zip(axes, valid_keys):
        res = results[gk]
        freqs  = res['sliding_freqs']
        times  = res['sliding_times']
        avg_tf = np.nanmean(np.array(res['sliding_coh']), axis=0)  # mean over subjects

        im = ax.imshow(avg_tf, aspect='auto', origin='lower',
                       extent=[times[0], times[-1], freqs[0], freqs[-1]],
                       cmap='viridis', vmin=vmin, vmax=vmax,
                       interpolation='bilinear')

        ax.set_title(gk, fontsize=11, fontweight='bold',
                     color=COLOURS.get(gk, 'black'))
        ax.set_xlabel('Time (s)')
        if ax is axes[0]:
            ax.set_ylabel('Frequency (Hz)')

        plt.colorbar(im, ax=ax, label='Coherence', shrink=0.85)

    fig.tight_layout()
    path = os.path.join(OUT_DIR, 'figure_02_sliding_heatmaps.png')
    fig.savefig(path)
    plt.close(fig)
    print(f"  Saved: {path}")


def plot_rolling_correlation(results):
    """Figure 3 — rolling Pearson correlation (Fisher-Z) over time."""
    fig, axes = plt.subplots(1, 2, figsize=(14, 5), sharey=True)
    fig.suptitle('Rolling Correlation: BNST–CEA\n'
                 f'(Pearson r, Fisher-Z, {ROLL_WIN_SEC} s window, '
                 f'{ROLL_STEP_SEC} s step)',
                 fontsize=13, fontweight='bold')

    panels = {
        axes[0]: ['Males PRE',   'Males POST'],
        axes[1]: ['Females PRE', 'Females POST'],
    }

    for ax, group_keys in panels.items():
        title_sex = group_keys[0].split()[0]
        ax.set_title(f'{title_sex}', fontsize=12)

        for gk in group_keys:
            if gk not in results or not results[gk]['rolling_r']:
                continue
            times = results[gk]['rolling_times']
            r_list = results[gk]['rolling_r']
            colour = COLOURS[gk]

            # Individual traces (thin, transparent)
            for r in r_list:
                ax.plot(times, r, color=colour, linewidth=0.7, alpha=0.30)

            # Mean ± SEM
            m, sem = mean_sem(r_list)
            ax.plot(times, m, color=colour, linewidth=2.0, label=gk)
            ax.fill_between(times, m - sem, m + sem,
                            color=colour, alpha=SHADE_ALPHA)

        ax.axhline(0, color='black', linewidth=0.8, linestyle='--', alpha=0.5)
        ax.set_xlabel('Time (s)')
        ax.set_ylabel('Correlation (Fisher Z)')
        ax.legend(framealpha=0.8, fontsize=9)

    fig.tight_layout()
    path = os.path.join(OUT_DIR, 'figure_03_rolling_correlation.png')
    fig.savefig(path)
    plt.close(fig)
    print(f"  Saved: {path}")


def plot_summary_bars(results):
    """Figure 4 — bar chart: mean coherence and mean |correlation| by group."""
    group_keys = ['Males PRE', 'Males POST', 'Females PRE', 'Females POST']
    valid_keys = [gk for gk in group_keys if gk in results]

    fig, axes = plt.subplots(1, 2, figsize=(11, 5))
    fig.suptitle('Mean Coherence & Correlation Summary\n(BNST–CEA)',
                 fontsize=13, fontweight='bold')

    x_pos  = np.arange(len(valid_keys))
    labels = valid_keys
    colours = [COLOURS[gk] for gk in valid_keys]

    # ── Mean coherence ────────────────────────────────────────────────────────
    ax = axes[0]
    means, sems = [], []
    for gk in valid_keys:
        vals = results[gk]['mean_coherence']
        means.append(np.nanmean(vals) if vals else np.nan)
        sems.append(np.nanstd(vals, ddof=1) / math.sqrt(len(vals)) if len(vals) > 1 else 0)

    bars = ax.bar(x_pos, means, color=colours, width=0.55,
                  edgecolor='black', linewidth=0.8, zorder=3)
    ax.errorbar(x_pos, means, yerr=sems, fmt='none',
                color='black', capsize=5, linewidth=1.5, zorder=4)

    # Individual data points
    for gi, gk in enumerate(valid_keys):
        vals = results[gk]['mean_coherence']
        jitter = np.random.default_rng(42).uniform(-0.12, 0.12, len(vals))
        ax.scatter(gi + jitter, vals, color='white', edgecolors='black',
                   s=30, zorder=5, linewidth=0.8)

    ax.set_xticks(x_pos)
    ax.set_xticklabels(labels, rotation=25, ha='right', fontsize=9)
    ax.set_ylabel(f'Mean Coherence ({FMIN}–{FMAX} Hz)')
    ax.set_title('Spectral Coherence')
    ax.set_ylim(bottom=0)

    # ── Mean |rolling correlation| ────────────────────────────────────────────
    ax = axes[1]
    means2, sems2 = [], []
    for gk in valid_keys:
        vals = results[gk]['mean_correlation']
        means2.append(np.nanmean(vals) if vals else np.nan)
        sems2.append(np.nanstd(vals, ddof=1) / math.sqrt(len(vals)) if len(vals) > 1 else 0)

    ax.bar(x_pos, means2, color=colours, width=0.55,
           edgecolor='black', linewidth=0.8, zorder=3)
    ax.errorbar(x_pos, means2, yerr=sems2, fmt='none',
                color='black', capsize=5, linewidth=1.5, zorder=4)

    for gi, gk in enumerate(valid_keys):
        vals = results[gk]['mean_correlation']
        jitter = np.random.default_rng(42).uniform(-0.12, 0.12, len(vals))
        ax.scatter(gi + jitter, vals, color='white', edgecolors='black',
                   s=30, zorder=5, linewidth=0.8)

    ax.set_xticks(x_pos)
    ax.set_xticklabels(labels, rotation=25, ha='right', fontsize=9)
    ax.set_ylabel('Mean |Correlation| (Fisher Z)')
    ax.set_title('Rolling Correlation')
    ax.set_ylim(bottom=0)

    fig.tight_layout()
    path = os.path.join(OUT_DIR, 'figure_04_summary_bars.png')
    fig.savefig(path)
    plt.close(fig)
    print(f"  Saved: {path}")

# ─────────────────────────────────────────────────────────────────────────────
# EXPORTS
# ─────────────────────────────────────────────────────────────────────────────

def export_summary_xlsx(results):
    """Export per-subject summary statistics to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Connectivity Summary'

    headers = ['Group', 'Subject', 'Analysis Type',
               'Mean Coherence (0.05–5 Hz)', 'Mean |Correlation| (Fisher Z)']
    ws.append(headers)

    for gk, res in results.items():
        subjects = res['subjects']
        for i, subj in enumerate(subjects):
            mean_coh  = res['mean_coherence'][i]  if i < len(res['mean_coherence'])  else ''
            mean_corr = res['mean_correlation'][i] if i < len(res['mean_correlation']) else ''

            if mean_coh != '':
                ws.append([gk, subj, 'Static Coherence',
                            round(mean_coh, 4), ''])
            if mean_corr != '':
                ws.append([gk, subj, 'Rolling Correlation',
                            '', round(mean_corr, 4)])

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 45)

    path = os.path.join(OUT_DIR, 'coherence_summary.xlsx')
    wb.save(path)
    print(f"  Saved: {path}")


def export_frequency_csv(results):
    """Export per-subject frequency-domain static coherence values."""
    path = os.path.join(OUT_DIR, 'coherence_frequency_data.csv')
    with open(path, 'w', newline='') as f:
        w = csv.writer(f)
        w.writerow(['Group', 'Subject', 'Frequency (Hz)', 'Coherence'])
        for gk, res in results.items():
            if not res['static_coh']:
                continue
            freqs = res['static_freqs']
            for i, subj in enumerate(res['subjects']):
                if i >= len(res['static_coh']):
                    continue
                coh = res['static_coh'][i]
                for freq, c in zip(freqs, coh):
                    w.writerow([gk, subj, round(float(freq), 4), round(float(c), 4)])
    print(f"  Saved: {path}")


def export_rolling_csv(results):
    """Export per-subject rolling-correlation time series."""
    path = os.path.join(OUT_DIR, 'coherence_rolling_data.csv')
    with open(path, 'w', newline='') as f:
        w = csv.writer(f)
        w.writerow(['Group', 'Subject', 'Time (s)', 'Correlation (Fisher Z)'])
        for gk, res in results.items():
            if not res['rolling_r']:
                continue
            times = res['rolling_times']
            for i, subj in enumerate(res['subjects']):
                if i >= len(res['rolling_r']):
                    continue
                r_vals = res['rolling_r'][i]
                for t, r in zip(times, r_vals):
                    if not math.isnan(r):
                        w.writerow([gk, subj, round(float(t), 3), round(float(r), 4)])
    print(f"  Saved: {path}")

# ─────────────────────────────────────────────────────────────────────────────
# EPOCH ANALYSIS  (pre-bout: t < 0  vs  post-bout onset: 0 ≤ t ≤ 10 s)
# ─────────────────────────────────────────────────────────────────────────────

# Use a shorter nperseg so both 30-s and 10-s epochs are valid
EPOCH_NPERSEG_SEC = 4    # s  →  ~80 samples at ~20 Hz; freq res = 0.25 Hz
EPOCH_FMIN        = 0.05
EPOCH_FMAX        = 5.0
EPOCH_POST_TMAX   = 10.0  # seconds after bout onset to include


def run_epoch_analysis():
    """Compute coherence separately for pre-bout (t<0) and post-bout (0–10 s)
    within each sex × condition group.  Returns a nested dict:

        epoch_results[group_key][epoch_label] = {
            'subjects', 'static_coh', 'static_freqs', 'mean_coherence'
        }
    """
    print("\nRunning epoch-based coherence analysis…")
    sex_data = {
        'Males':   load_sex_data(MALES_XLSX),
        'Females': load_sex_data(FEMALES_XLSX),
    }

    epoch_results = {}

    for sex, sex_dict in sex_data.items():
        for cond, cond_dict in sex_dict.items():
            group_key = f'{sex} {cond}'
            print(f"  {group_key}…")

            subjects  = cond_dict['BNST'][0]
            time_vec  = cond_dict['BNST'][1]
            bnst_data = cond_dict['BNST'][2]
            cea_data  = cond_dict['CEA'][2]

            fs = infer_fs(time_vec)

            pre_mask  = time_vec < 0
            post_mask = (time_vec >= 0) & (time_vec <= EPOCH_POST_TMAX)

            epoch_results[group_key] = {}

            for epoch_label, mask in [('Pre-bout (t<0)', pre_mask),
                                       (f'Post-bout (0–{EPOCH_POST_TMAX:.0f}s)', post_mask)]:
                n_samp = int(mask.sum())
                print(f"    {epoch_label}: {n_samp} samples ({n_samp/fs:.1f} s)")

                er = {
                    'subjects':      subjects,
                    'static_freqs':  None,
                    'static_coh':    [],
                    'mean_coherence': [],
                    'n_samples':     n_samp,
                    'fs':            fs,
                }

                for i, subj in enumerate(subjects):
                    s1 = bnst_data[mask, i]
                    s2 = cea_data[mask, i]

                    f, c = calculate_static_coherence(
                        s1, s2, fs,
                        fmin=EPOCH_FMIN,
                        fmax=EPOCH_FMAX,
                        nperseg_sec=EPOCH_NPERSEG_SEC,
                    )
                    if c is not None:
                        if er['static_freqs'] is None:
                            er['static_freqs'] = f
                        er['static_coh'].append(c)
                        er['mean_coherence'].append(float(np.mean(c)))

                epoch_results[group_key][epoch_label] = er

    return epoch_results


# ──────────────────────────────────────────────────────────────────────────────
# EPOCH FIGURES
# ──────────────────────────────────────────────────────────────────────────────

def plot_epoch_coherence_spectra(epoch_results):
    """Figure 5 — coherence spectra per epoch (pre vs post) in each group.

    Layout: 2 rows (PRE condition, POST condition) × 2 cols (Males, Females).
    Within each panel both epochs are overlaid.
    """
    conds = ['PRE', 'POST']
    sexes = ['Males', 'Females']

    epoch_labels = None
    for gk, epochs in epoch_results.items():
        if epochs:
            epoch_labels = list(epochs.keys())
            break
    if epoch_labels is None:
        return

    epoch_colours  = ['#2176AE', '#D62246']   # blue = pre-bout, red = post-bout
    epoch_linestyle = ['-', '--']

    fig, axes = plt.subplots(len(conds), len(sexes),
                              figsize=(12, 9), sharex=True, sharey=True)
    fig.suptitle('Epoch Coherence: BNST–CEA\n'
                 f'Pre-bout (t<0) vs Post-bout onset (0–{EPOCH_POST_TMAX:.0f}s)  '
                 f'[nperseg={EPOCH_NPERSEG_SEC}s, {EPOCH_FMIN}–{EPOCH_FMAX}Hz]',
                 fontsize=13, fontweight='bold')

    for ri, cond in enumerate(conds):
        for ci, sex in enumerate(sexes):
            ax  = axes[ri][ci]
            gk  = f'{sex} {cond}'
            ax.set_title(f'{sex} — {cond}', fontsize=11,
                         color=COLOURS.get(f'{sex} {cond}', 'black'), fontweight='bold')

            if gk not in epoch_results:
                continue

            for ep_idx, ep_label in enumerate(epoch_labels):
                er = epoch_results[gk].get(ep_label)
                if er is None or not er['static_coh']:
                    continue

                freqs    = er['static_freqs']
                coh_list = er['static_coh']
                col      = epoch_colours[ep_idx]
                ls       = epoch_linestyle[ep_idx]

                # Individual traces
                for c in coh_list:
                    ax.plot(freqs, c, color=col, linewidth=0.7,
                            alpha=0.30, linestyle=ls)

                # Mean ± SEM
                m, sem = mean_sem(coh_list)
                ax.plot(freqs, m, color=col, linewidth=2.2,
                        linestyle=ls, label=ep_label)
                ax.fill_between(freqs, m - sem, m + sem,
                                color=col, alpha=SHADE_ALPHA)

            ax.set_ylim([0, 1])
            ax.legend(fontsize=8, framealpha=0.8)
            if ri == len(conds) - 1:
                ax.set_xlabel('Frequency (Hz)')
            if ci == 0:
                ax.set_ylabel('Coherence')

    fig.tight_layout()
    path = os.path.join(OUT_DIR, 'figure_05_epoch_coherence_spectra.png')
    fig.savefig(path)
    plt.close(fig)
    print(f"  Saved: {path}")


def plot_epoch_summary_bars(epoch_results):
    """Figure 6 — grouped bar chart: mean coherence per epoch across all groups."""
    group_keys = ['Males PRE', 'Males POST', 'Females PRE', 'Females POST']

    epoch_labels = None
    for gk, epochs in epoch_results.items():
        if epochs:
            epoch_labels = list(epochs.keys())
            break
    if epoch_labels is None:
        return

    n_groups = len(group_keys)
    n_epochs = len(epoch_labels)
    bar_w    = 0.35
    offsets  = np.linspace(-(n_epochs - 1) * bar_w / 2,
                            (n_epochs - 1) * bar_w / 2, n_epochs)
    epoch_colours = ['#2176AE', '#D62246']

    fig, ax = plt.subplots(figsize=(11, 6))
    fig.suptitle('Mean Coherence by Epoch: BNST–CEA\n'
                 f'Pre-bout (t<0) vs Post-bout onset (0–{EPOCH_POST_TMAX:.0f}s)',
                 fontsize=13, fontweight='bold')

    x_pos = np.arange(n_groups)

    for ep_idx, ep_label in enumerate(epoch_labels):
        means, sems, all_vals = [], [], []
        for gk in group_keys:
            er = epoch_results.get(gk, {}).get(ep_label)
            vals = er['mean_coherence'] if er else []
            means.append(np.nanmean(vals) if vals else np.nan)
            sems.append(np.nanstd(vals, ddof=1) / math.sqrt(len(vals))
                        if len(vals) > 1 else 0)
            all_vals.append(vals)

        col = epoch_colours[ep_idx]
        xb  = x_pos + offsets[ep_idx]
        ax.bar(xb, means, width=bar_w, color=col, label=ep_label,
               edgecolor='black', linewidth=0.8, alpha=0.85, zorder=3)
        ax.errorbar(xb, means, yerr=sems, fmt='none',
                    color='black', capsize=4, linewidth=1.5, zorder=4)

        rng = np.random.default_rng(42 + ep_idx)
        for gi, vals in enumerate(all_vals):
            if vals:
                jitter = rng.uniform(-bar_w * 0.3, bar_w * 0.3, len(vals))
                ax.scatter(xb[gi] + jitter, vals,
                           color='white', edgecolors='black',
                           s=28, zorder=5, linewidth=0.8)

    ax.set_xticks(x_pos)
    ax.set_xticklabels(group_keys, rotation=20, ha='right', fontsize=10)
    ax.set_ylabel(f'Mean Coherence ({EPOCH_FMIN}–{EPOCH_FMAX} Hz)')
    ax.set_ylim(bottom=0)
    ax.legend(fontsize=10, framealpha=0.8)

    fig.tight_layout()
    path = os.path.join(OUT_DIR, 'figure_06_epoch_summary_bars.png')
    fig.savefig(path)
    plt.close(fig)
    print(f"  Saved: {path}")


# ──────────────────────────────────────────────────────────────────────────────
# EPOCH EXPORTS
# ──────────────────────────────────────────────────────────────────────────────

def export_epoch_summary_xlsx(epoch_results):
    """Per-subject mean coherence for each epoch, exported to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Epoch Coherence Summary'

    headers = ['Group', 'Subject',
               'Epoch', 'N Samples', 'Duration (s)',
               f'Mean Coherence ({EPOCH_FMIN}–{EPOCH_FMAX} Hz)']
    ws.append(headers)

    for gk, epochs in epoch_results.items():
        for ep_label, er in epochs.items():
            fs   = er['fs']
            n_s  = er['n_samples']
            dur  = round(n_s / fs, 2)
            for i, subj in enumerate(er['subjects']):
                mc = er['mean_coherence'][i] if i < len(er['mean_coherence']) else ''
                ws.append([gk, subj, ep_label, n_s, dur,
                            round(mc, 4) if mc != '' else ''])

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 45)

    path = os.path.join(OUT_DIR, 'epoch_coherence_summary.xlsx')
    wb.save(path)
    print(f"  Saved: {path}")


def export_epoch_frequency_csv(epoch_results):
    """Per-subject, per-epoch frequency-domain coherence values."""
    path = os.path.join(OUT_DIR, 'epoch_coherence_frequency_data.csv')
    with open(path, 'w', newline='') as f:
        w = csv.writer(f)
        w.writerow(['Group', 'Epoch', 'Subject', 'Frequency (Hz)', 'Coherence'])
        for gk, epochs in epoch_results.items():
            for ep_label, er in epochs.items():
                if not er['static_coh']:
                    continue
                freqs = er['static_freqs']
                for i, subj in enumerate(er['subjects']):
                    if i >= len(er['static_coh']):
                        continue
                    coh = er['static_coh'][i]
                    for freq, c in zip(freqs, coh):
                        w.writerow([gk, ep_label, subj,
                                    round(float(freq), 4), round(float(c), 4)])
    print(f"  Saved: {path}")


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    run_analysis()

    # Epoch analysis
    epoch_results = run_epoch_analysis()
    print("\nGenerating epoch figures…")
    plot_epoch_coherence_spectra(epoch_results)
    plot_epoch_summary_bars(epoch_results)
    print("\nExporting epoch data…")
    export_epoch_summary_xlsx(epoch_results)
    export_epoch_frequency_csv(epoch_results)
    print(f"\nEpoch outputs saved to:\n  {OUT_DIR}")
